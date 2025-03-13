"""
This is a python class file to read and write data from the excel file

NOTE: Install openpyxl before using this file
"""
from openpyxl import load_workbook

class ExcelFunctions:

    def __init__(self, file_name, sheet_name):
        self.file = file_name # file path
        self.sheet = sheet_name # exact sheet number

    # method to get the row count
    def row_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_row
    
    # method to get the column count
    def column_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_column
    
    # method to read data from the excel file
    def read_data(self, row_number, column_number):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.cell(row=row_number, column=column_number).value

    # method to write the data into the excel file
    def write_data(self, row_number, column_number, data):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        sheet.cell(row=row_number, column=column_number).value = data
        workbook.save(self.file)